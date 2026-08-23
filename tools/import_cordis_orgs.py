#!/usr/bin/env python3
"""Turn the CORDIS participants table into a loadable organisation seed.

    tools/import_cordis_orgs.py PARTICIPANTS.{xlsx,csv} -o seed/003_cordis_organisations.sql

CORDIS publishes, for every EU-funded project, the organisations that signed
for it: a PIC, a legal name, a country, an activity type. That is a register
of who does battery research in Europe, and it is exactly what
`bd.organization` is for.

**Scope is the important argument here.** An organisation's own record is
official EC data, but whether it belongs in a battery database depends on
which projects you counted, and that classification is a heuristic. The
workbook's `BATTERY_INTEGRATED` class means "a wider project that contains a
battery workstream" - wide enough that its participant list includes a
consortium for the protection of Pecorino Toscano cheese, swept in from an
agricultural-digitalisation project. Two thirds of the 6,353 participants
arrive that way.

So the default is `--scope strict`: only participants of projects the
workbook classed `BATTERY_CORE` or `BATTERY_ECOSYSTEM`, where batteries are
the primary object. That is 2,042 organisations, each of which signed a
grant whose subject is actually batteries. `--scope all` restores the full
list for anyone who wants the wider net and knows what is in it.

What this script will *not* do is import the rest of the CORDIS record. EC
contribution, project counts and SME status are facts about a funding
programme, not about an organisation, and this schema has nowhere honest to
put them. They stay in CORDIS.

Four normalisations happen here, each one a place where the source and this
schema genuinely disagree:

  country       CORDIS uses the EU's own codes. `UK` and `EL` are not ISO
                3166-1 alpha-2; `organization.country` says it is. Mapped.

  roles         CORDIS activity type says whether an organisation is a
                university, a research institute or a company. It does not
                say whether it makes cells. So HES/REC become {lab} and
                everything else stays empty rather than guessed at. An empty
                role is recoverable; a wrong one is not.

  uid           Slugged from the legal name, because `org/fraunhofer-...`
                is greppable and `org/pic-999984059` is not. Nine slugs
                collide across 6,353 organisations; those get their PIC
                appended, which is the only tiebreak guaranteed to be stable.

  quoting       The export was CSV-quoted twice, so a name containing
                quotation marks arrives with them multiplied - the Greek
                research centre Demokritos ends up wrapped in two opening
                and four closing quote characters. Runs of quotes collapse
                back to one.

Note what is *not* written: no `organization_alias` rows. The only two names
CORDIS gives are the legal name and the short name, and both already have a
column. An alias would be the same string a second time under a `kind` that
would have to be invented.

The source file is not committed - it is a 20 MB EC bulk export. Fetch it
from CORDIS, run this, commit the SQL.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

# CORDIS uses the European Commission's country codes, which differ from
# ISO 3166-1 alpha-2 in exactly these two places.
EU_TO_ISO = {"UK": "GB", "EL": "GR"}

# Activity types that describe a research performer. 'HSE' is a transposition
# of 'HES' present in the source; the full-text variants appear on records
# that came through the Funding & Tenders Portal rather than the bulk files.
RESEARCH_ACTIVITY = {
    "HES",
    "HSE",
    "REC",
    "Higher education",
    "Higher or secondary education establishment",
    "Research organisation",
}

MAX_SLUG = 60


def unquote(text: str) -> str:
    """Undo the source's double-applied CSV quoting."""
    return re.sub(r'"{2,}', '"', text).strip()


def normalise_url(url: str) -> str:
    """CORDIS mixes `https://x.com`, `www.x.com` and `x.com` in one column."""
    url = url.strip()
    if not url:
        return ""
    if not re.match(r"^https?://", url, re.I):
        url = "https://" + url.lstrip("/")
    return url


def slugify(text: str) -> str:
    """Slug, cut at a word boundary rather than mid-syllable."""
    ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    slug = re.sub(r"-+", "-", re.sub(r"[^A-Za-z0-9]+", "-", ascii_text).strip("-").lower())
    if len(slug) <= MAX_SLUG:
        return slug
    cut = slug[:MAX_SLUG]
    # Prefer the last complete word; fall back to a hard cut for a single
    # very long token, which is rare enough not to matter.
    return (cut.rsplit("-", 1)[0] if "-" in cut else cut).strip("-")


# Tokens that must not be title-cased: legal forms that are conventionally
# capitalised a particular way, and anything that is an acronym rather than
# a word. Everything else in an all-caps name is a shouted word.
LEGAL_FORMS = {
    "bv": "BV", "nv": "NV", "sa": "SA", "srl": "SRL", "sl": "SL", "sro": "SRO",
    "spa": "SpA", "gmbh": "GmbH", "ag": "AG", "kg": "KG", "ev": "eV", "as": "AS",
    "ab": "AB", "oy": "OY", "aps": "ApS", "plc": "PLC", "ltd": "Ltd", "llc": "LLC",
    "inc": "Inc", "sas": "SAS", "sarl": "SARL", "scpa": "ScpA", "zoo": "ZOO",
    "doo": "DOO", "ou": "OU", "uab": "UAB", "sia": "SIA", "kft": "Kft", "sp": "SP",
    "vzw": "VZW", "asbl": "ASBL", "aisbl": "AISBL", "sca": "SCA", "sce": "SCE",
    "cic": "CIC", "llp": "LLP", "pte": "Pte", "pty": "Pty", "bhd": "Bhd",
}

# Small words that stay lowercase inside a name, unless they lead it.
MINOR_WORDS = {
    "and", "of", "the", "for", "in", "on", "at", "to", "de", "du", "des", "da",
    "di", "del", "della", "der", "die", "das", "den", "van", "von", "y", "e",
    "et", "la", "le", "les", "el", "il", "och", "og", "voor", "en", "za", "na",
    # German, Dutch, Nordic and Slavic connectives that are short enough to
    # be mistaken for acronyms. DECHEMA's registered name contains "FUR".
    "fur", "und", "mit", "zur", "zum", "auf", "aus", "bei", "im", "am", "ter",
    "ve", "og", "pa", "av", "til", "og", "i", "u", "w", "z", "s", "o", "a",
    "per", "con", "com", "dos", "das", "dei", "ed", "og", "es", "si",
}


def display_name(name: str) -> str:
    """CORDIS shouts. `legal_name` keeps the shout; `name` is for reading.

    Only all-caps strings are touched, and only word-shaped tokens within
    them: an acronym, an initialism or anything carrying a digit or an
    internal full stop is left exactly as the Commission wrote it.
    """
    if not name or not name.isupper():
        return name

    def fix(token: str, first: bool) -> str:
        bare = re.sub(r"[^A-Za-z]", "", token)
        if not bare:
            return token
        low = bare.lower()
        if low in LEGAL_FORMS:
            return token.replace(bare, LEGAL_FORMS[low])
        # Acronyms and initialisms: too short to be a word, or punctuated
        # like one (R.FLO, S.P.A.), or carrying digits.
        if len(bare) <= 3 and low not in MINOR_WORDS:
            return token
        if re.search(r"[0-9]", token) or re.search(r"[A-Z]\.[A-Z]", token):
            return token
        if low in MINOR_WORDS and not first:
            return token.replace(bare, low)
        return token.replace(bare, bare.capitalize())

    tokens = name.split()
    return " ".join(fix(t, i == 0) for i, t in enumerate(tokens))


def normalise_for_match(name: str) -> str:
    ascii_name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    stripped = re.sub(r"[^a-z0-9 ]", " ", ascii_name.lower())
    words = [w for w in stripped.split() if w not in LEGAL_FORMS]
    return " ".join(words)


def find_duplicate_groups(records: list[dict]) -> list[list[dict]]:
    """Organisations that look like the same body under two PICs.

    Not merged - CORDIS genuinely issues more than one PIC to some bodies,
    and separately, unrelated organisations share a name across borders.
    Telling those apart is a human's job, so this only reports.
    """
    groups: dict[str, list[dict]] = {}
    for record in records:
        key = normalise_for_match(record["short"] or record["legal"])
        if key:
            groups.setdefault(key, []).append(record)
    return [sorted(g, key=lambda r: r["pic"]) for g in groups.values() if len(g) > 1]


def sql_str(value: str | None) -> str:
    """A SQL literal. NULL for anything that is not a real value."""
    if value is None:
        return "NULL"
    value = value.strip()
    if not value:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def read_rows(path: Path, sheet_hint: str = "participant") -> list[dict]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            import openpyxl
        except ImportError:  # pragma: no cover - dependency hint
            sys.exit("reading .xlsx needs openpyxl; pip install openpyxl (or pass a CSV)")
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = next(
            (book[name] for name in book.sheetnames if sheet_hint in name.lower()),
            None,
        )
        if sheet is None:
            sys.exit(f"no {sheet_hint} sheet in {path.name}: {book.sheetnames}")
        rows = sheet.iter_rows(values_only=True)
        header = ["" if c is None else str(c).strip() for c in next(rows)]
        return [
            dict(zip(header, ["" if c is None else str(c).strip() for c in row]))
            for row in rows
            if any(c is not None and str(c).strip() for c in row)
        ]

    csv.field_size_limit(10**9)
    with path.open(encoding="utf-8", newline="") as handle:
        return [
            {k: (v or "").strip() for k, v in row.items()}
            for row in csv.DictReader(handle)
        ]


def is_name_like(text: str) -> bool:
    """Reject a 'short name' that is really a registration number.

    One participant has its German VAT number, DE29 581 94 16, in the
    short-name field while the legal name reads EVTEC GmbH. A name has
    more letters than digits.
    """
    letters = sum(c.isalpha() for c in text)
    digits = sum(c.isdigit() for c in text)
    return letters > digits and letters >= 2


def pick(row: dict, *names: str) -> str:
    """First non-empty value among column names that may vary by export."""
    for name in names:
        if row.get(name):
            return row[name]
    return ""


# Scope classes where batteries are the primary object of the project, as
# opposed to BATTERY_INTEGRATED, which only requires a battery workstream
# inside a project about something else.
STRICT_SCOPES = {"BATTERY_CORE", "BATTERY_ECOSYSTEM"}


def strict_project_ids(rows: list[dict]) -> set[str]:
    ids = set()
    for row in rows:
        pid = pick(row, "Project ID", "projectID", "id")
        if pid and pick(row, "Scope Class", "scopeClass") in STRICT_SCOPES:
            ids.add(pid)
    return ids


def collapse(rows: list[dict], keep_projects: set[str] | None = None) -> list[dict]:
    """One record per PIC, preferring the most complete spelling of each field."""
    by_pic: dict[str, dict] = {}
    for row in rows:
        if keep_projects is not None:
            pid = pick(row, "Project ID", "projectID")
            if pid not in keep_projects:
                continue
        pic = pick(row, "PIC", "Organisation ID / PIC", "organisationID")
        if not pic or not pic.isdigit():
            continue
        legal = unquote(pick(row, "Legal Name", "name", "legalName"))
        if not legal:
            continue
        record = by_pic.setdefault(
            pic, {"pic": pic, "legal": "", "short": "", "country": "", "url": "", "activity": ""}
        )
        # Longest spelling wins: CORDIS truncates names in some exports.
        if len(legal) > len(record["legal"]):
            record["legal"] = legal
        for key, columns in (
            ("short", ("Short Name", "shortName")),
            ("country", ("Country", "country")),
            ("url", ("Organisation URL", "organizationURL")),
            ("activity", ("Activity Type", "activityType")),
        ):
            value = unquote(pick(row, *columns))
            if key == "short" and value and not is_name_like(value):
                continue
            if value and not record[key]:
                record[key] = value
    return sorted(by_pic.values(), key=lambda r: int(r["pic"]))


def assign_uids(records: list[dict]) -> None:
    """Slug from the name a person would use, not the one on the charter.

    Trinity College Dublin is registered as "The Provost, Fellows,
    Foundation Scholars and the other members of Board...", which slugs to
    something no one will ever type or recognise. The short name is the
    better identifier wherever CORDIS supplies one; the legal name is the
    fallback, and the PIC breaks the ties.
    """
    def base(record: dict) -> str:
        return slugify(record["short"]) or slugify(record["legal"])

    counts = Counter(base(r) for r in records)
    for record in records:
        slug = base(record) or f"pic-{record['pic']}"
        if counts[slug] > 1:
            slug = f"{slug}-pic{record['pic']}"
        record["uid"] = f"org/{slug}"


def build_sql(records: list[dict], source_name: str, retrieved: str, scope: str) -> str:
    stats: Counter = Counter()
    values: list[str] = []

    for record in records:
        country = EU_TO_ISO.get(record["country"], record["country"])
        if len(country) != 2:
            country = ""
            stats["country dropped"] += 1
        elif record["country"] in EU_TO_ISO:
            stats["country remapped"] += 1

        roles = "{lab}" if record["activity"] in RESEARCH_ACTIVITY else "{}"
        stats["role lab" if roles == "{lab}" else "role unknown"] += 1

        # `name` is what a person would write; `legal_name` is what the
        # grant was signed as, kept verbatim. They are often not the same
        # string, and CORDIS writes both in capitals.
        display = display_name(record["short"] or record["legal"])
        if display != (record["short"] or record["legal"]):
            stats["name un-shouted"] += 1
        values.append(
            "  ({uid}, {name}, {legal}, {country}, '{roles}', {url}, {pic})".format(
                uid=sql_str(record["uid"]),
                name=sql_str(display),
                legal=sql_str(record["legal"]),
                country=sql_str(country),
                roles=roles,
                url=sql_str(normalise_url(record["url"])),
                pic=sql_str(record["pic"]),
            )
        )
        if normalise_url(record["url"]) != record["url"].strip():
            stats["url scheme added"] += 1

    scope_note = (
        "signed at least one EU grant whose primary subject is batteries\n"
        "-- (CORDIS scope BATTERY_CORE or BATTERY_ECOSYSTEM)"
        if scope == "strict"
        else "appear in any project the source workbook classed as\n"
        "-- battery-related, including wider projects with a battery workstream"
    )
    header = f"""-- =====================================================================
-- battery-data : seed/003_cordis_organisations.sql
--
-- {len(records):,} organisations that {scope_note}.
-- From the CORDIS participants register.
--
-- GENERATED FILE - do not edit by hand.
--   tools/import_cordis_orgs.py {source_name} --scope {scope}
--   source retrieved {retrieved}
--
-- Scope matters more than it looks. Each organisation record below is
-- official EC data, but *which* organisations are listed depends on a
-- heuristic classification of projects. Widening to BATTERY_INTEGRATED
-- triples the count and admits participants of projects that merely
-- contain a battery workstream - a cheese consortium among them.
--
-- These are identities, not claims. A row here says "the European
-- Commission knows this organisation by this name, in this country, under
-- this PIC" - nothing about what it manufactures, measures or publishes.
-- Nothing in this file is an observation, and nothing in it needs a
-- provenance record, because no value here is a measurement.
--
-- roles is {{lab}} only where CORDIS classified the participant as a
-- university or a research institute. Companies are left with no role:
-- CORDIS says a firm took EU research money, not that it makes cells.
--
-- No organization_alias rows: CORDIS gives a legal name and a short name,
-- and both already have a column of their own.
--
-- Re-runnable, and composes with 001 and 002 in any order. Where an
-- organisation is already present as curated data, only empty columns are
-- filled: a curated name, role or country is never overwritten by CORDIS.
-- =====================================================================

SET search_path = bd, public;

-- ---------------------------------------------------------------------
-- Organisations
-- ---------------------------------------------------------------------
INSERT INTO organization (uid, name, legal_name, country, roles, website, pic) VALUES
"""

    # An organisation may already be here as curated data - Audi arrives via
    # seed/002 with roles={manufacturer}. DO NOTHING would drop its PIC on
    # the floor, so instead fill the holes and touch nothing that was set
    # deliberately: name, roles and country stay exactly as curated.
    body = ",\n".join(values) + """
ON CONFLICT (uid) DO UPDATE SET
      pic        = COALESCE(organization.pic, EXCLUDED.pic),
      legal_name = COALESCE(organization.legal_name, EXCLUDED.legal_name),
      website    = COALESCE(organization.website, EXCLUDED.website),
      updated_at = now()
  WHERE organization.pic IS NULL
     OR organization.legal_name IS NULL
     OR organization.website IS NULL;
"""

    duplicate_note = ""
    groups = find_duplicate_groups(records)
    if groups:
        lines = []
        for group in sorted(groups, key=lambda g: g[0]["legal"]):
            names = " | ".join(
                f"{r['pic']} {EU_TO_ISO.get(r['country'], r['country']) or '??'}"
                for r in group
            )
            lines.append(f"--   {group[0]['short'] or group[0]['legal']}: {names}")
        duplicate_note = (
            "\n-- ---------------------------------------------------------------------\n"
            f"-- {len(groups)} name collisions across {sum(len(g) for g in groups)} rows, listed\n"
            "-- rather than merged. Some are one body holding two PICs; others are\n"
            "-- unrelated organisations that share a name across two countries.\n"
            "-- Telling those apart is a human's job.\n"
            "-- ---------------------------------------------------------------------\n"
            + "\n".join(lines)
            + "\n"
        )

    footer = duplicate_note + f"""
-- ---------------------------------------------------------------------
-- Counts, so a failed load is obvious rather than silent.
-- ---------------------------------------------------------------------
DO $$
DECLARE n bigint;
BEGIN
  SELECT count(*) INTO n FROM organization WHERE pic IS NOT NULL;
  RAISE NOTICE 'CORDIS organisations present: %', n;
END $$;
"""
    return header + body + footer, stats


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("source", type=Path, help="CORDIS participants .xlsx or .csv")
    parser.add_argument("-o", "--output", type=Path, default=Path("seed/003_cordis_organisations.sql"))
    parser.add_argument("--retrieved", default="", help="retrieval date of the export (YYYY-MM-DD)")
    parser.add_argument(
        "--scope",
        choices=("strict", "all"),
        default="strict",
        help="strict (default): only participants of BATTERY_CORE/ECOSYSTEM projects",
    )
    parser.add_argument(
        "--projects",
        type=Path,
        help="projects table, if the participants were passed as a bare CSV",
    )
    args = parser.parse_args()

    if not args.source.exists():
        sys.exit(f"no such file: {args.source}")

    rows = read_rows(args.source)

    keep = None
    if args.scope == "strict":
        project_source = args.projects or args.source
        if args.projects is None and args.source.suffix.lower() not in {".xlsx", ".xlsm"}:
            sys.exit("--scope strict needs the projects table too; pass --projects, or --scope all")
        keep = strict_project_ids(read_rows(project_source, sheet_hint="project"))
        if not keep:
            sys.exit("no strict-scope projects found - check the projects table")

    records = collapse(rows, keep)
    if not records:
        sys.exit("no organisation rows found - is this the participants table?")
    assign_uids(records)

    retrieved = args.retrieved or dt.date.today().isoformat()
    sql, stats = build_sql(records, args.source.name, retrieved, args.scope)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(sql, encoding="utf-8")

    print(f"{args.source.name}: {len(rows):,} participant rows, scope={args.scope}")
    if keep is not None:
        print(f"  kept participants of {len(keep):,} strict-scope projects")
    print(f"  -> {len(records):,} organisations, {len(set(r['uid'] for r in records)):,} unique uids")
    for key in sorted(stats):
        print(f"     {stats[key]:>6,}  {key}")
    groups = find_duplicate_groups(records)
    if groups:
        print(f"     {sum(len(g) for g in groups):>6,}  rows in {len(groups)} name collisions (listed in the seed, not merged)")
    print(f"  -> {args.output} ({args.output.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
